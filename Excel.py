import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
import sys
import time

def getXLfile():
    print("Files Found:")
    files = glob.glob("*.xlsx")
    for i in range(0, len(files)):
        print("[%d] %s" % (i, files[i]))

    UserInputValid = False

    while(not UserInputValid):
        userinput = input("Please Select File To Process (Enter Assigned Number):")
        try:
            temp = int(userinput)
            if temp in range(0, len(files)):
                os.system('cls')
                print(files[temp], "was selected\n")
                UserInputValid = True
            else:
                print("Please Enter Number within Range of Files\n")
        except ValueError:
            print("Please Enter a Number\n")
    return files[temp]


def writeExcel(outQ, wb, sheet, filename):
    cell_title = sheet['O1']
    cell_title.value = "Room"
    cell_title.font = Font(bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    cell_title = sheet['P1']
    cell_title.value = "Instructor"
    cell_title.font = Font(bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    while not outQ.empty():
        i = outQ.get()
        sheet.cell(row=i[0], column=15).value = i[1]
        sheet.cell(row=i[0], column=16).value = i[2]

    wb.save(filename + ".xlsx")


def loadexcel(name):
    data = []
    try:
        wb = load_workbook(name)
    except:
        print("Please rename excel document to data.xlsx")
        time.sleep(10)
        sys.exit()

    first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(first_sheet)
    cell_title = worksheet['O1']
    cell_title.value = "Room"
    cell_title.font = Font(bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(2, worksheet.max_row + 1):
        id = (worksheet.cell(row=row, column=1).value)
        data.append([row, id])
    return data, wb, worksheet

