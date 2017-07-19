from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import sys
import time
from multiprocessing import Queue


header = {'User-Agent': 'Mozilla/5.0 (X11; OpenBSD i386) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.125 Safari/537.36}'}

def Minput():

    print("******************************************************************************")
    print("Intructions:\n1) Generate Class Screening Report from Careworks")
    print("Excel file has following format:")
    print("ClassId  Title   Category    location    Address .....\n")
    print("2) Save generated report as a xlsx file within Excel\n")
    print("3) Place new file in the same location as this program file\n")
    print("4) Enter the excel file name\n")
    print("5) Wait for program to finish, new excel will be created in same location as\nprogram file\n")
    print("******************************************************************************")

    name = input("Please enter the excel filename (Exclude .xlsx): ")
    return name

def grabRoom(id, header):

    try:
        print("Adding Room Location for clinic id ", id, end="....", flush=True)
        url = "https://www.inova.org/creg/classdetails.aspx?sid=1&ClassID=%d&sslRedirect=true"
        htmlContent = requests.get(url % id, headers=header, timeout=60, allow_redirects=True)
        soup = BeautifulSoup(htmlContent.text, 'html.parser')
        text = soup.find(text="Room: ")
        tag = text.parent
        room = tag.next_sibling

    except:
        room = "Error"

    return room


try:
    filename = Minput()
    wb = load_workbook(filename+'.xlsx')
except:
    print("Please enter correct name! Be sure not to include extra spaces")
    print("Program will auto close in.....")
    print("3")
    time.sleep(1)
    print("2")
    time.sleep(1)
    print("1")
    time.sleep(1)
    sys.exit()

first_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(first_sheet)
cell_title = worksheet['O1']
cell_title.value = "Room"
cell_title.font = Font(bold=True)
cell_title.alignment = Alignment(horizontal="center", vertical="center")
for row in range(2, worksheet.max_row):
    for column in "A":
        cell_name = "{}{}".format(column,row)
        id = worksheet[cell_name].value
        if id == None:
            continue
        room = grabRoom(id, header)
        worksheet.cell(row=row, column=15).value = room
        wb.save(filename+'-Updated.xlsx')
        print("Success!\n")

print("Program finished! Will close in....")
print("3")
time.sleep(1)
print("2")
time.sleep(1)
print("1")
time.sleep(1)
sys.exit()